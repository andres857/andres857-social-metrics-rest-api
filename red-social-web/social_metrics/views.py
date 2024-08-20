from django.shortcuts import render
from django.http import HttpResponse
import pandas as pd
from datetime import datetime
import openpyxl

def get_all_institutions(request):
    return HttpResponse("Hello, world. Institutions")

def uploadFile(request):
    return render(request, 'uploadfile.html')

def index(request):

    if "GET" == request.method:
        return render(request, 'uploadfile.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        return render(request, 'uploadfile.html', {"excel_data":excel_data})
    

def convertir_nombre_pestana_a_fecha(nombre_pestana):
    # Asumiendo que el nombre de la pestaña está en formato 'YYYY-MM'
    return datetime.strptime(nombre_pestana, '%Y-%m')

def procesar_datos_excel(request):
    if "GET" == request.method:
        return render(request, 'uploadfile.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        xls = pd.ExcelFile(excel_file)
        
        for nombre_pestana in xls.sheet_names:
            fecha_recoleccion = convertir_nombre_pestana_a_fecha(nombre_pestana)
            print(f"\nProcesando datos para el período: {nombre_pestana}")
            print(f"Fecha de recolección: {fecha_recoleccion}")
            
            # Leer datos de la pestaña
            df = pd.read_excel(xls, sheet_name=nombre_pestana)
            
            print("\nDatos de la pestaña:")
            print(df)
            
            print("\nEstadísticas resumidas:")
            print(df.describe())
            
            # print("\nCálculo de engagement rate para cada fila:")
            # for _, row in df.iterrows():
            #     engagement_rate = calcular_engagement_rate(row['likes'], row['seguidores'])
            #     print(f"Institución ID: {row['institucion_id']}, Red Social ID: {row['red_social_id']}, Engagement Rate: {engagement_rate:.2f}%")

def calcular_engagement_rate(likes, seguidores):
    return (likes / seguidores * 100) if seguidores > 0 else 0
